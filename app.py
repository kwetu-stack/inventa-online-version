import os
import sqlite3
import io
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import qrcode
from openpyxl import Workbook, load_workbook

# ---------- APP CONFIG ----------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "supersecret")  # For Render, override via environment
TENANT = 'digitalclub'
DB_PATH = 'inventory.db'

# ---------- HELPER ----------
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def is_logged_in():
    return 'user_id' in session

# ---------- LOGIN & AUTH ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE username = ? AND password = ?",
            (username, password)
        ).fetchone()
        conn.close()

        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash("Login successful!", "info")
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid username or password.", "error")
            return redirect(url_for('login'))
    return render_template('login.html', tenant=TENANT)

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for('login'))

@app.route(f'/{TENANT}/change-password', methods=['GET', 'POST'])
def change_password():
    if not is_logged_in():
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],)).fetchone()

        if user and user['password'] == current_password:
            conn.execute("UPDATE users SET password = ? WHERE id = ?", (new_password, session['user_id']))
            conn.commit()
            flash("Password updated successfully!", "info")
        else:
            flash("Current password is incorrect.", "error")
        conn.close()
        return redirect(url_for('dashboard'))

    return render_template('change-password.html', tenant=TENANT)

# ---------- DASHBOARD ----------
@app.route('/')
def home():
    return redirect(f'/{TENANT}/dashboard')

@app.route(f'/{TENANT}/dashboard')
def dashboard():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products').fetchall()
    total_products = len(products)
    total_sales = conn.execute('SELECT SUM(grand_total) FROM sales').fetchone()[0] or 0
    sales_data = conn.execute('SELECT date, SUM(grand_total) as total FROM sales GROUP BY date ORDER BY date').fetchall()
    conn.close()

    dates = [row['date'] for row in sales_data]
    totals = [row['total'] for row in sales_data]

    return render_template('dashboard.html',
                           tenant=TENANT,
                           total_products=total_products,
                           total_sales=total_sales,
                           dates=dates,
                           totals=totals)

# ---------- VIEW PRODUCTS ----------
@app.route(f'/{TENANT}/products')
def view_products():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products').fetchall()
    conn.close()
    return render_template('view-products.html', products=products, tenant=TENANT)

# ---------- ADD STOCK ----------
@app.route(f'/{TENANT}/add-stock', methods=['GET', 'POST'])
def add_stock():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    if request.method == 'POST':
        product_id = int(request.form['product_id'])
        quantity = int(request.form['quantity'])

        conn.execute('UPDATE products SET quantity = quantity + ? WHERE id = ?', (quantity, product_id))
        conn.commit()
        conn.close()
        flash("Stock added successfully.", "info")
        return redirect(f'/{TENANT}/add-stock')

    products = conn.execute('SELECT id, name FROM products').fetchall()
    conn.close()
    return render_template('add-stock.html', products=products, tenant=TENANT)

# ---------- RECORD SALE ----------
@app.route(f'/{TENANT}/record-sale', methods=['GET', 'POST'])
def record_sale():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products').fetchall()

    if request.method == 'POST':
        customer = request.form['customer']
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')

        total_amount = 0
        selected_items = []

        for i, product_id in enumerate(product_ids):
            qty = int(quantities[i])
            if qty > 0:
                product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
                total = qty * product['price']
                total_amount += total
                selected_items.append((product_id, qty, product['price']))

        vat = round(total_amount * 0.16, 2)
        grand_total = round(total_amount + vat, 2)

        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO sales (customer, total_amount, vat, grand_total, date) VALUES (?, ?, ?, ?, DATE("now"))',
            (customer, total_amount, vat, grand_total)
        )
        sale_id = cursor.lastrowid

        for product_id, qty, price in selected_items:
            cursor.execute('INSERT INTO sale_items (sale_id, product_id, quantity, price) VALUES (?, ?, ?, ?)',
                           (sale_id, product_id, qty, price))
            cursor.execute('UPDATE products SET quantity = quantity - ? WHERE id = ?', (qty, product_id))

        conn.commit()
        conn.close()
        flash("Sale recorded successfully!", "info")
        return redirect(f'/{TENANT}/sales-history')

    conn.close()
    return render_template('record-sale.html', products=products, tenant=TENANT)

# ---------- SALES HISTORY ----------
@app.route(f'/{TENANT}/sales-history')
def sales_history():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    sales = conn.execute('SELECT * FROM sales ORDER BY date DESC').fetchall()
    conn.close()
    return render_template('sales-history.html', sales=sales, tenant=TENANT)

# ---------- INVOICE ----------
@app.route(f'/{TENANT}/invoice/<int:sale_id>')
def download_invoice(sale_id):
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    sale = conn.execute('SELECT * FROM sales WHERE id = ?', (sale_id,)).fetchone()
    items = conn.execute('''
        SELECT p.name, si.quantity, si.price
        FROM sale_items si
        JOIN products p ON si.product_id = p.id
        WHERE si.sale_id = ?
    ''', (sale_id,)).fetchall()
    conn.close()
    return render_template('invoice.html', sale=sale, items=items, tenant=TENANT)

# ---------- QR ----------
@app.route(f'/{TENANT}/qr/<int:sale_id>')
def generate_qr(sale_id):
    if not is_logged_in():
        return redirect(url_for('login'))

    qr = qrcode.make(f"Invoice ID: {sale_id} | Tenant: {TENANT}")
    img_io = io.BytesIO()
    qr.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')

# ---------- INVENTORY TOOLS ----------
@app.route(f'/{TENANT}/inventory-tools')
def inventory_tools():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template('inventory-tools.html', tenant=TENANT)

# ---------- DOWNLOAD INVENTORY ----------
@app.route(f'/{TENANT}/download-inventory')
def download_inventory():
    if not is_logged_in():
        return redirect(url_for('login'))

    conn = get_db_connection()
    products = conn.execute('SELECT id, name, quantity, buying_price, price FROM products').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(['ID', 'Item', 'Stock', 'Buying Price (KES)', 'Selling Price (KES)'])

    for product in products:
        ws.append([product['id'], product['name'], product['quantity'], product['buying_price'], product['price']])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f"{TENANT}_inventory.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ---------- UPLOAD INVENTORY ----------
@app.route(f'/{TENANT}/upload-inventory', methods=['POST'])
def upload_inventory():
    if not is_logged_in():
        return redirect(url_for('login'))

    if 'file' not in request.files or request.files['file'].filename == '':
        flash("No file selected.", "error")
        return redirect(f'/{TENANT}/inventory-tools')

    file = request.files['file']
    wb = load_workbook(file)
    ws = wb.active

    conn = get_db_connection()
    for row in ws.iter_rows(min_row=2, values_only=True):
        product_id, name, stock, buying_price, selling_price = row
        conn.execute("""
            UPDATE products
            SET name = ?, quantity = ?, buying_price = ?, price = ?
            WHERE id = ?
        """, (name, stock, buying_price, selling_price, product_id))
    conn.commit()
    conn.close()

    flash("Inventory updated successfully!", "info")
    return redirect(f'/{TENANT}/inventory-tools')

# ---------- ERROR HANDLERS ----------
@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

if __name__ == '__main__':
    app.run(debug=True)
